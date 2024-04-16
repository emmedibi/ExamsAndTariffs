from excelCreation.constants import *
from excelCreation.classes import *
from openpyxl import Workbook
from openpyxl import load_workbook

## FILE THAT WORKS ON EXCEL FILE:
## - READ THE TARIFF FILE DATARULE FROM A PRE-EXISTED EXCEL FILE
## - CREATE A NEW EXCEL FILE AND APPLIED THE RULES FOR EACH TARIFF

def extractTariffesInformation(sheet):
    tariffes = []

    for col in sheet.iter_rows(min_row=2, values_only=True):
        tariff = Tariff(id=col[ID_TARIFFARIO],
                        name=col[NOME_TARIFFARIO],
                        discountPercentage=col[REGOLA_DI_SCONTO])
        tariffes.append(tariff)
    return tariffes

def calculatePrice(examfullprice, discountPercentage):
    return examfullprice*discountPercentage


def createExcelFileForNewExam(exam, tariffes):
    wb = Workbook()
    sheet = wb.sheetnames
    sheet = wb.active
    sheet.append(["ID_PRESTAZIONE", "NOME_PRESTAZIONE", "ID_TARIFFARIO", "NOME_TARIFFARIO", "TARIFF"])
    for t in tariffes :
        priceOfT = calculatePrice(exam.price, t.discountPercentage)
        examTariffForExcel = [exam.id, exam.name, t.id, t.name, priceOfT]
        sheet.append(examTariffForExcel)
    wb.save(filename='Test_' + exam.id + '.xlsx')
    return "Test_" + exam.id + ".xlsx"

def createDataForExcel(idE, name, price, tariffName):
    workbook = load_workbook(filename="sampleTariff.xlsx", read_only=True)
    sheetname = "Privato"
    if tariffName in workbook.sheetnames:
        sheet = workbook[tariffName]
    else :
        return ValueError("Please provide valid input")
    
    print(f"Worksheet names: {workbook.sheetnames}")
    print(f"The title of the Worksheet is: {sheet.title}")

    exam = ExamData(id=idE,
                name=name,
                price=price,
                tariff_type=tariffName)
    print(exam)
    tariffes = []
    tariffes = extractTariffesInformation(sheet)
    print("######### TARIFFES OUTPUT: ")
    print(tariffes)
    return createExcelFileForNewExam(exam, tariffes)